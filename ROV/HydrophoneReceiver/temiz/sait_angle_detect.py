from multiprocessing import Process
import multiprocessing
import time
import serial
#import RPi.GPIO as gpio
import numpy as np
import matplotlib.pyplot as plt
import math as m
import pandas as pd
import os
import openpyxl

MODE_SELECT_AUTO = "1".encode('utf-8')
MODE_SELECT_MANUAL = "2".encode('utf-8')

IFFT_ENABLED = "A".encode('utf-8')
IFFT_DISABLED = "B".encode('utf-8')

DESCRIBE_YOURSELF_FLAG = "0".encode('utf-8')

SEARCH_1SEC = "9".encode('utf-8')
SEARCH_2SEC = "8".encode('utf-8')
SEARCH_3SEC = "7".encode('utf-8')
SEARCH_4SEC = "6".encode('utf-8')
SEARCH_5SEC = "5".encode('utf-8')

BANDWIDTH_3 = "Z".encode('utf-8')
BANDWIDTH_5 = "Y".encode('utf-8')
BANDWIDTH_7 = "X".encode('utf-8')
BANDWIDTH_9 = "W".encode('utf-8')
BANDWIDTH_11 = "V".encode('utf-8')
BANDWIDTH_13 = "U".encode('utf-8')

# setting up the trigger.
EXTERNAL_TRIGGER_PIN = 20
#gpio.setmode(gpio.BCM)
#gpio.setup(EXTERNAL_TRIGGER_PIN, gpio.OUT)

q1, q2, q3, q4 = multiprocessing.Queue(), multiprocessing.Queue(), multiprocessing.Queue(), multiprocessing.Queue() # Queues for processes.

USBports = [serial.Serial(), serial.Serial(), serial.Serial(), serial.Serial()]
def DetectWhichPort():
    global USBports
    for i in range(0, 4):
        try:
            port = serial.Serial("/dev/ttyACM"+str(i), baudrate=115200, timeout=0)
            port.write(DESCRIBE_YOURSELF_FLAG)
            data = port.readline()
            while data == b'':
                data = port.readline()
            queue = int(data.decode('utf-8'))
            USBports[queue-1] = port
        except:
            None

def Config(port):
    port.write(MODE_SELECT_AUTO)
    time.sleep(0.1)
    port.write(IFFT_DISABLED)
    time.sleep(0.1)
    port.write(SEARCH_2SEC)
    time.sleep(0.1)
    port.write(BANDWIDTH_3)
    time.sleep(0.1)

def MyProc1(port):
    Config(port)
    #gpio.output(EXTERNAL_TRIGGER_PIN, gpio.HIGH)
    global q1
    while 1:
        try:
            data = port.readline()
            while data == b'':
                data = port.readline()
            q1.put(data.decode('utf-8'))
            #gpio.output(EXTERNAL_TRIGGER_PIN, gpio.LOW)
        except:
            None

def MyProc2(port):
    Config(port)
    #gpio.output(EXTERNAL_TRIGGER_PIN, gpio.HIGH)
    global q2
    while 1:
        try:
            data = port.readline()
            while data == b'':
                data = port.readline()
            q2.put(data.decode('utf-8'))
            #gpio.output(EXTERNAL_TRIGGER_PIN, gpio.LOW)
        except:
            None

def MyProc3(port):
    Config(port)
    #gpio.output(EXTERNAL_TRIGGER_PIN, gpio.HIGH)
    global q3
    while 1:
        try:
            data = port.readline()
            while data == b'':
                data = port.readline()
            q3.put(data.decode('utf-8'))
            #gpio.output(EXTERNAL_TRIGGER_PIN, gpio.LOW)
        except:
            None

def MyProc4(port):
    Config(port)
    #gpio.output(EXTERNAL_TRIGGER_PIN, gpio.HIGH)
    global q4
    while 1:
        try:
            data = port.readline()
            while data == b'':
                data = port.readline()
            q4.put(data.decode('utf-8'))
            #gpio.output(EXTERNAL_TRIGGER_PIN, gpio.LOW)
        except:
            None












def data_preparetion(input_matrix):
    print("Starting to data preparation!\n")

    while True:


#        print("2. On derece")
#        print("3. Eksi on derece")
#        print("0. Çıkış")


        sel = int(input("Please enter the degree: "))


        if sel == -180:
            max_number = 0
            if not os.path.exists("data/-180"):
                os.mkdir("data/-180")

            df = pd.DataFrame(input_matrix)
            file_names = os.listdir("data/-180")

            if not file_names:
                file_name='0.xlsx'

            else:
                for name in file_names:
                    if name.endswith(".xlsx"):
                        number = int(name[:-5])
                        if number > max_number:
                            max_number = number
                file_name = str(max_number+1) + ".xlsx"

            writer = pd.ExcelWriter('data/-180/'+file_name, engine='xlsxwriter')
            df.to_excel(writer, sheet_name='Sheet1', index=False, header=False)
            writer._save()

        elif sel == -170:
            max_number = 0
            if not os.path.exists("data/-170"):
                os.mkdir("data/-180")

            df = pd.DataFrame(input_matrix)
            file_names = os.listdir("data/-170")

            if not file_names:
                file_name = '0.xlsx'

            else:
                for name in file_names:
                    if name.endswith(".xlsx"):
                        number = int(name[:-5])
                        if number > max_number:
                            max_number = number
                file_name = str(max_number + 1) + ".xlsx"

            writer = pd.ExcelWriter('data/-180/' + file_name, engine='xlsxwriter')
            df.to_excel(writer, sheet_name='Sheet1', index=False, header=False)

            writer._save()

        elif sel == -160:
            max_number = 0
            if not os.path.exists("data/-160"):
                os.mkdir("data/-160")

            df = pd.DataFrame(input_matrix)
            file_names = os.listdir("data/-160")

            if not file_names:
                file_name='0.xlsx'

            else:
                for name in file_names:
                    if name.endswith(".xlsx"):
                        number = int(name[:-5])
                        if number > max_number:
                            max_number = number
                file_name = str(max_number+1) + ".xlsx"

            writer = pd.ExcelWriter('data/-160/'+file_name, engine='xlsxwriter')
            df.to_excel(writer, sheet_name='Sheet1', index=False, header=False)
            writer._save()

        elif sel == -150:
            max_number = 0
            if not os.path.exists("data/-150"):
                os.mkdir("data/-150")

            df = pd.DataFrame(input_matrix)
            file_names = os.listdir("data/-150")

            if not file_names:
                file_name = '0.xlsx'

            else:
                for name in file_names:
                    if name.endswith(".xlsx"):
                        number = int(name[:-5])
                        if number > max_number:
                            max_number = number
                file_name = str(max_number + 1) + ".xlsx"

            writer = pd.ExcelWriter('data/-150/' + file_name, engine='xlsxwriter')
            df.to_excel(writer, sheet_name='Sheet1', index=False, header=False)
            writer._save()

        elif sel == -140:
            max_number = 0
            if not os.path.exists("data/-140"):
                os.mkdir("data/-140")

            df = pd.DataFrame(input_matrix)
            file_names = os.listdir("data/-140")

            if not file_names:
                file_name = '0.xlsx'

            else:
                for name in file_names:
                    if name.endswith(".xlsx"):
                        number = int(name[:-5])
                        if number > max_number:
                            max_number = number
                file_name = str(max_number + 1) + ".xlsx"

            writer = pd.ExcelWriter('data/-140/' + file_name, engine='xlsxwriter')
            df.to_excel(writer, sheet_name='Sheet1', index=False, header=False)
            writer._save()

        elif sel == -130:
            max_number = 0
            if not os.path.exists("data/-130"):
                os.mkdir("data/-130")

            df = pd.DataFrame(input_matrix)
            file_names = os.listdir("data/-130")

            if not file_names:
                file_name = '0.xlsx'

            else:
                for name in file_names:
                    if name.endswith(".xlsx"):
                        number = int(name[:-5])
                        if number > max_number:
                            max_number = number
                file_name = str(max_number + 1) + ".xlsx"

            writer = pd.ExcelWriter('data/-130/' + file_name, engine='xlsxwriter')
            df.to_excel(writer, sheet_name='Sheet1', index=False, header=False)
            writer._save()

        elif sel == -120:
            max_number = 0
            if not os.path.exists("data/-120"):
                os.mkdir("data/-120")

            df = pd.DataFrame(input_matrix)
            file_names = os.listdir("data/-120")

            if not file_names:
                file_name = '0.xlsx'

            else:
                for name in file_names:
                    if name.endswith(".xlsx"):
                        number = int(name[:-5])
                        if number > max_number:
                            max_number = number
                file_name = str(max_number + 1) + ".xlsx"

            writer = pd.ExcelWriter('data/-120/' + file_name, engine='xlsxwriter')
            df.to_excel(writer, sheet_name='Sheet1', index=False, header=False)
            writer._save()

        elif sel == -110:
            max_number = 0
            if not os.path.exists("data/-110"):
                os.mkdir("data/-110")

            df = pd.DataFrame(input_matrix)
            file_names = os.listdir("data/-110")

            if not file_names:
                file_name = '0.xlsx'

            else:
                for name in file_names:
                    if name.endswith(".xlsx"):
                        number = int(name[:-5])
                        if number > max_number:
                            max_number = number
                file_name = str(max_number + 1) + ".xlsx"

            writer = pd.ExcelWriter('data/-110/' + file_name, engine='xlsxwriter')
            df.to_excel(writer, sheet_name='Sheet1', index=False, header=False)
            writer._save()

        elif sel == -100:
            max_number = 0
            if not os.path.exists("data/-100"):
                os.mkdir("data/-100")

            df = pd.DataFrame(input_matrix)
            file_names = os.listdir("data/-100")

            if not file_names:
                file_name = '0.xlsx'

            else:
                for name in file_names:
                    if name.endswith(".xlsx"):
                        number = int(name[:-5])
                        if number > max_number:
                            max_number = number
                file_name = str(max_number + 1) + ".xlsx"

            writer = pd.ExcelWriter('data/-100/' + file_name, engine='xlsxwriter')
            df.to_excel(writer, sheet_name='Sheet1', index=False, header=False)
            writer._save()

        elif sel == -90:
            max_number = 0
            if not os.path.exists("data/-90"):
                os.mkdir("data/-90")

            df = pd.DataFrame(input_matrix)
            file_names = os.listdir("data/-90")

            if not file_names:
                file_name = '0.xlsx'

            else:
                for name in file_names:
                    if name.endswith(".xlsx"):
                        number = int(name[:-5])
                        if number > max_number:
                            max_number = number
                file_name = str(max_number + 1) + ".xlsx"

            writer = pd.ExcelWriter('data/-90/' + file_name, engine='xlsxwriter')
            df.to_excel(writer, sheet_name='Sheet1', index=False, header=False)
            writer._save()

        elif sel == -80:
            max_number = 0
            if not os.path.exists("data/-80"):
                os.mkdir("data/-80")

            df = pd.DataFrame(input_matrix)
            file_names = os.listdir("data/-80")

            if not file_names:
                file_name = '0.xlsx'

            else:
                for name in file_names:
                    if name.endswith(".xlsx"):
                        number = int(name[:-5])
                        if number > max_number:
                            max_number = number
                file_name = str(max_number + 1) + ".xlsx"

            writer = pd.ExcelWriter('data/-80/' + file_name, engine='xlsxwriter')
            df.to_excel(writer, sheet_name='Sheet1', index=False, header=False)
            writer._save()

        elif sel == -70:
            max_number = 0
            if not os.path.exists("data/-70"):
                os.mkdir("data/-70")

            df = pd.DataFrame(input_matrix)
            file_names = os.listdir("data/-70")

            if not file_names:
                file_name = '0.xlsx'

            else:
                for name in file_names:
                    if name.endswith(".xlsx"):
                        number = int(name[:-5])
                        if number > max_number:
                            max_number = number
                file_name = str(max_number + 1) + ".xlsx"

            writer = pd.ExcelWriter('data/-70/' + file_name, engine='xlsxwriter')
            df.to_excel(writer, sheet_name='Sheet1', index=False, header=False)
            writer._save()

        elif sel == -60:
            max_number = 0
            if not os.path.exists("data/-60"):
                os.mkdir("data/-60")

            df = pd.DataFrame(input_matrix)
            file_names = os.listdir("data/-60")

            if not file_names:
                file_name = '0.xlsx'

            else:
                for name in file_names:
                    if name.endswith(".xlsx"):
                        number = int(name[:-5])
                        if number > max_number:
                            max_number = number
                file_name = str(max_number + 1) + ".xlsx"

            writer = pd.ExcelWriter('data/-60/' + file_name, engine='xlsxwriter')
            df.to_excel(writer, sheet_name='Sheet1', index=False, header=False)
            writer._save()

        elif sel == -50:
            max_number = 0
            if not os.path.exists("data/-50"):
                os.mkdir("data/-50")

            df = pd.DataFrame(input_matrix)
            file_names = os.listdir("data/-50")

            if not file_names:
                file_name = '0.xlsx'

            else:
                for name in file_names:
                    if name.endswith(".xlsx"):
                        number = int(name[:-5])
                        if number > max_number:
                            max_number = number
                file_name = str(max_number + 1) + ".xlsx"

            writer = pd.ExcelWriter('data/-60/' + file_name, engine='xlsxwriter')
            df.to_excel(writer, sheet_name='Sheet1', index=False, header=False)
            writer._save()

        elif sel == -40:
            max_number = 0
            if not os.path.exists("data/-40"):
                os.mkdir("data/-40")

            df = pd.DataFrame(input_matrix)
            file_names = os.listdir("data/-40")

            if not file_names:
                file_name = '0.xlsx'

            else:
                for name in file_names:
                    if name.endswith(".xlsx"):
                        number = int(name[:-5])
                        if number > max_number:
                            max_number = number
                file_name = str(max_number + 1) + ".xlsx"

            writer = pd.ExcelWriter('data/-40/' + file_name, engine='xlsxwriter')
            df.to_excel(writer, sheet_name='Sheet1', index=False, header=False)
            writer._save()

        elif sel == -30:
            max_number = 0
            if not os.path.exists("data/-30"):
                os.mkdir("data/-30")

            df = pd.DataFrame(input_matrix)
            file_names = os.listdir("data/-30")

            if not file_names:
                file_name = '0.xlsx'

            else:
                for name in file_names:
                    if name.endswith(".xlsx"):
                        number = int(name[:-5])
                        if number > max_number:
                            max_number = number
                file_name = str(max_number + 1) + ".xlsx"

            writer = pd.ExcelWriter('data/-30/' + file_name, engine='xlsxwriter')
            df.to_excel(writer, sheet_name='Sheet1', index=False, header=False)
            writer._save()

        elif sel == -20:
            max_number = 0
            if not os.path.exists("data/-20"):
                os.mkdir("data/-20")

            df = pd.DataFrame(input_matrix)
            file_names = os.listdir("data/-20")

            if not file_names:
                file_name = '0.xlsx'

            else:
                for name in file_names:
                    if name.endswith(".xlsx"):
                        number = int(name[:-5])
                        if number > max_number:
                            max_number = number
                file_name = str(max_number + 1) + ".xlsx"

            writer = pd.ExcelWriter('data/-20/' + file_name, engine='xlsxwriter')
            df.to_excel(writer, sheet_name='Sheet1', index=False, header=False)
            writer._save()

        elif sel == -10:
            max_number = 0
            if not os.path.exists("data/-10"):
                os.mkdir("data/-10")

            df = pd.DataFrame(input_matrix)
            file_names = os.listdir("data/-10")

            if not file_names:
                file_name = '0.xlsx'

            else:
                for name in file_names:
                    if name.endswith(".xlsx"):
                        number = int(name[:-5])
                        if number > max_number:
                            max_number = number
                file_name = str(max_number + 1) + ".xlsx"

            writer = pd.ExcelWriter('data/-10/' + file_name, engine='xlsxwriter')
            df.to_excel(writer, sheet_name='Sheet1', index=False, header=False)
            writer._save()

        elif sel == 0:
            max_number = 0
            if not os.path.exists("data/0"):
                os.mkdir("data/0")

            df = pd.DataFrame(input_matrix)
            file_names = os.listdir("data/0")

            if not file_names:
                file_name = '0.xlsx'

            else:
                for name in file_names:
                    if name.endswith(".xlsx"):
                        number = int(name[:-5])
                        if number > max_number:
                            max_number = number
                file_name = str(max_number + 1) + ".xlsx"

            writer = pd.ExcelWriter('data/0/' + file_name, engine='xlsxwriter')
            df.to_excel(writer, sheet_name='Sheet1', index=False, header=False)
            writer._save()

        elif sel == 180:
            max_number = 0
            if not os.path.exists("data/180"):
                os.mkdir("data/180")

            df = pd.DataFrame(input_matrix)
            file_names = os.listdir("data/180")

            if not file_names:
                file_name = '0.xlsx'

            else:
                for name in file_names:
                    if name.endswith(".xlsx"):
                        number = int(name[:-5])
                        if number > max_number:
                            max_number = number
                file_name = str(max_number + 1) + ".xlsx"

            writer = pd.ExcelWriter('data/180/' + file_name, engine='xlsxwriter')
            df.to_excel(writer, sheet_name='Sheet1', index=False, header=False)
            writer._save()

        elif sel == 170:
            max_number = 0
            if not os.path.exists("data/170"):
                os.mkdir("data/170")

            df = pd.DataFrame(input_matrix)
            file_names = os.listdir("data/170")

            if not file_names:
                file_name = '0.xlsx'

            else:
                for name in file_names:
                    if name.endswith(".xlsx"):
                        number = int(name[:-5])
                        if number > max_number:
                            max_number = number
                file_name = str(max_number + 1) + ".xlsx"

            writer = pd.ExcelWriter('data/170/' + file_name, engine='xlsxwriter')
            df.to_excel(writer, sheet_name='Sheet1', index=False, header=False)

            writer._save()

        elif sel == 160:
            max_number = 0
            if not os.path.exists("data/160"):
                os.mkdir("data/160")

            df = pd.DataFrame(input_matrix)
            file_names = os.listdir("data/160")

            if not file_names:
                file_name = '0.xlsx'

            else:
                for name in file_names:
                    if name.endswith(".xlsx"):
                        number = int(name[:-5])
                        if number > max_number:
                            max_number = number
                file_name = str(max_number + 1) + ".xlsx"

            writer = pd.ExcelWriter('data/160/' + file_name, engine='xlsxwriter')
            df.to_excel(writer, sheet_name='Sheet1', index=False, header=False)
            writer._save()

        elif sel == 150:
            max_number = 0
            if not os.path.exists("data/150"):
                os.mkdir("data/150")

            df = pd.DataFrame(input_matrix)
            file_names = os.listdir("data/150")

            if not file_names:
                file_name = '0.xlsx'

            else:
                for name in file_names:
                    if name.endswith(".xlsx"):
                        number = int(name[:-5])
                        if number > max_number:
                            max_number = number
                file_name = str(max_number + 1) + ".xlsx"

            writer = pd.ExcelWriter('data/150/' + file_name, engine='xlsxwriter')
            df.to_excel(writer, sheet_name='Sheet1', index=False, header=False)
            writer._save()

        elif sel == 140:
            max_number = 0
            if not os.path.exists("data/140"):
                os.mkdir("data/140")

            df = pd.DataFrame(input_matrix)
            file_names = os.listdir("data/140")

            if not file_names:
                file_name = '0.xlsx'

            else:
                for name in file_names:
                    if name.endswith(".xlsx"):
                        number = int(name[:-5])
                        if number > max_number:
                            max_number = number
                file_name = str(max_number + 1) + ".xlsx"

            writer = pd.ExcelWriter('data/140/' + file_name, engine='xlsxwriter')
            df.to_excel(writer, sheet_name='Sheet1', index=False, header=False)
            writer._save()

        elif sel == 130:
            max_number = 0
            if not os.path.exists("data/130"):
                os.mkdir("data/130")

            df = pd.DataFrame(input_matrix)
            file_names = os.listdir("data/130")

            if not file_names:
                file_name = '0.xlsx'

            else:
                for name in file_names:
                    if name.endswith(".xlsx"):
                        number = int(name[:-5])
                        if number > max_number:
                            max_number = number
                file_name = str(max_number + 1) + ".xlsx"

            writer = pd.ExcelWriter('data/130/' + file_name, engine='xlsxwriter')
            df.to_excel(writer, sheet_name='Sheet1', index=False, header=False)
            writer._save()

        elif sel == 120:
            max_number = 0
            if not os.path.exists("data/120"):
                os.mkdir("data/120")

            df = pd.DataFrame(input_matrix)
            file_names = os.listdir("data/120")

            if not file_names:
                file_name = '0.xlsx'

            else:
                for name in file_names:
                    if name.endswith(".xlsx"):
                        number = int(name[:-5])
                        if number > max_number:
                            max_number = number
                file_name = str(max_number + 1) + ".xlsx"

            writer = pd.ExcelWriter('data/120/' + file_name, engine='xlsxwriter')
            df.to_excel(writer, sheet_name='Sheet1', index=False, header=False)
            writer._save()

        elif sel == 110:
            max_number = 0
            if not os.path.exists("data/110"):
                os.mkdir("data/110")

            df = pd.DataFrame(input_matrix)
            file_names = os.listdir("data/110")

            if not file_names:
                file_name = '0.xlsx'

            else:
                for name in file_names:
                    if name.endswith(".xlsx"):
                        number = int(name[:-5])
                        if number > max_number:
                            max_number = number
                file_name = str(max_number + 1) + ".xlsx"

            writer = pd.ExcelWriter('data/110/' + file_name, engine='xlsxwriter')
            df.to_excel(writer, sheet_name='Sheet1', index=False, header=False)
            writer._save()

        elif sel == 100:
            max_number = 0
            if not os.path.exists("data/100"):
                os.mkdir("data/100")

            df = pd.DataFrame(input_matrix)
            file_names = os.listdir("data/100")

            if not file_names:
                file_name = '0.xlsx'

            else:
                for name in file_names:
                    if name.endswith(".xlsx"):
                        number = int(name[:-5])
                        if number > max_number:
                            max_number = number
                file_name = str(max_number + 1) + ".xlsx"

            writer = pd.ExcelWriter('data/100/' + file_name, engine='xlsxwriter')
            df.to_excel(writer, sheet_name='Sheet1', index=False, header=False)
            writer._save()

        elif sel == 90:
            max_number = 0
            if not os.path.exists("data/90"):
                os.mkdir("data/90")

            df = pd.DataFrame(input_matrix)
            file_names = os.listdir("data/90")

            if not file_names:
                file_name = '0.xlsx'

            else:
                for name in file_names:
                    if name.endswith(".xlsx"):
                        number = int(name[:-5])
                        if number > max_number:
                            max_number = number
                file_name = str(max_number + 1) + ".xlsx"

            writer = pd.ExcelWriter('data/90/' + file_name, engine='xlsxwriter')
            df.to_excel(writer, sheet_name='Sheet1', index=False, header=False)
            writer._save()

        elif sel == 80:
            max_number = 0
            if not os.path.exists("data/80"):
                os.mkdir("data/80")

            df = pd.DataFrame(input_matrix)
            file_names = os.listdir("data/80")

            if not file_names:
                file_name = '0.xlsx'

            else:
                for name in file_names:
                    if name.endswith(".xlsx"):
                        number = int(name[:-5])
                        if number > max_number:
                            max_number = number
                file_name = str(max_number + 1) + ".xlsx"

            writer = pd.ExcelWriter('data/80/' + file_name, engine='xlsxwriter')
            df.to_excel(writer, sheet_name='Sheet1', index=False, header=False)
            writer._save()

        elif sel == 70:
            max_number = 0
            if not os.path.exists("data/70"):
                os.mkdir("data/70")

            df = pd.DataFrame(input_matrix)
            file_names = os.listdir("data/70")

            if not file_names:
                file_name = '0.xlsx'

            else:
                for name in file_names:
                    if name.endswith(".xlsx"):
                        number = int(name[:-5])
                        if number > max_number:
                            max_number = number
                file_name = str(max_number + 1) + ".xlsx"

            writer = pd.ExcelWriter('data/70/' + file_name, engine='xlsxwriter')
            df.to_excel(writer, sheet_name='Sheet1', index=False, header=False)
            writer._save()

        elif sel == 60:
            max_number = 0
            if not os.path.exists("data/60"):
                os.mkdir("data/60")

            df = pd.DataFrame(input_matrix)
            file_names = os.listdir("data/60")

            if not file_names:
                file_name = '0.xlsx'

            else:
                for name in file_names:
                    if name.endswith(".xlsx"):
                        number = int(name[:-5])
                        if number > max_number:
                            max_number = number
                file_name = str(max_number + 1) + ".xlsx"

            writer = pd.ExcelWriter('data/60/' + file_name, engine='xlsxwriter')
            df.to_excel(writer, sheet_name='Sheet1', index=False, header=False)
            writer._save()

        elif sel == 50:
            max_number = 0
            if not os.path.exists("data/50"):
                os.mkdir("data/50")

            df = pd.DataFrame(input_matrix)
            file_names = os.listdir("data/50")

            if not file_names:
                file_name = '0.xlsx'

            else:
                for name in file_names:
                    if name.endswith(".xlsx"):
                        number = int(name[:-5])
                        if number > max_number:
                            max_number = number
                file_name = str(max_number + 1) + ".xlsx"

            writer = pd.ExcelWriter('data/50/' + file_name, engine='xlsxwriter')
            df.to_excel(writer, sheet_name='Sheet1', index=False, header=False)
            writer._save()

        elif sel == 40:
            max_number = 0
            if not os.path.exists("data/40"):
                os.mkdir("data/40")

            df = pd.DataFrame(input_matrix)
            file_names = os.listdir("data/40")

            if not file_names:
                file_name = '0.xlsx'

            else:
                for name in file_names:
                    if name.endswith(".xlsx"):
                        number = int(name[:-5])
                        if number > max_number:
                            max_number = number
                file_name = str(max_number + 1) + ".xlsx"

            writer = pd.ExcelWriter('data/40/' + file_name, engine='xlsxwriter')
            df.to_excel(writer, sheet_name='Sheet1', index=False, header=False)
            writer._save()

        elif sel == 30:
            max_number = 0
            if not os.path.exists("data/30"):
                os.mkdir("data/30")

            df = pd.DataFrame(input_matrix)
            file_names = os.listdir("data/30")

            if not file_names:
                file_name = '0.xlsx'

            else:
                for name in file_names:
                    if name.endswith(".xlsx"):
                        number = int(name[:-5])
                        if number > max_number:
                            max_number = number
                file_name = str(max_number + 1) + ".xlsx"

            writer = pd.ExcelWriter('data/30/' + file_name, engine='xlsxwriter')
            df.to_excel(writer, sheet_name='Sheet1', index=False, header=False)
            writer._save()

        elif sel == 20:
            max_number = 0
            if not os.path.exists("data/20"):
                os.mkdir("data/20")

            df = pd.DataFrame(input_matrix)
            file_names = os.listdir("data/20")

            if not file_names:
                file_name = '0.xlsx'

            else:
                for name in file_names:
                    if name.endswith(".xlsx"):
                        number = int(name[:-5])
                        if number > max_number:
                            max_number = number
                file_name = str(max_number + 1) + ".xlsx"

            writer = pd.ExcelWriter('data/20/' + file_name, engine='xlsxwriter')
            df.to_excel(writer, sheet_name='Sheet1', index=False, header=False)
            writer._save()

        elif sel == 10:
            max_number = 0
            if not os.path.exists("data/10"):
                os.mkdir("data/10")

            df = pd.DataFrame(input_matrix)
            file_names = os.listdir("data/10")

            if not file_names:
                file_name = '0.xlsx'

            else:
                for name in file_names:
                    if name.endswith(".xlsx"):
                        number = int(name[:-5])
                        if number > max_number:
                            max_number = number
                file_name = str(max_number + 1) + ".xlsx"

            writer = pd.ExcelWriter('data/10/' + file_name, engine='xlsxwriter')
            df.to_excel(writer, sheet_name='Sheet1', index=False, header=False)
            writer._save()

        elif sel == 52:
            print("Programdan çıkılıyor...")
            break
        else:
            print("Geçersiz seçim. Lütfen tekrar deneyin.")

def calculate_average():
    if not os.path.exists("data/results"):
        os.mkdir("data/results")

    matrix_size=(4,141)
    for i in range(len(os.listdir('data'))):
        dir_index=-180
        file_index=0
        sum=np.zeros(matrix_size)
        dir="data/"+str(dir_index)
        file_names=os.listdir(dir)
        len=len(file_names)

        for j in range(len):
            file_name=str(file_index)+".xlsx"
            wb=openpyxl.load_workbook("data/"+str(dir_index)+"/"+file_name)
            sheet=wb.active
            temp=[]
            for row in sheet.iter_rows(min_row=1,max_row=4,min_col=1,max_col=141,values_only=True):
                temp.append(row)
            array=np.array(temp)
            sum=array+sum

        average=sum/len


        df = pd.DataFrame(average)
        writer = pd.ExcelWriter("data/results/"+str(dir_index)+".xlsx", engine='xlsxwriter')
        df.to_excel(writer, sheet_name='Sheet1', index=False, header=False)
        writer._save()

        dir_index=dir_index+10

    print("---Average matrices calculates and saved into 'data/results'---")



    """
    fig, axs = plt.subplots(1, 3, figsize=(10, 4))

    axs[0].imshow(average_1,cmap='viridis',interpolation='nearest', aspect='auto')
    axs[1].imshow(average_2,cmap='viridis',interpolation='nearest', aspect='auto')
    axs[2].imshow(average_3,cmap='viridis',interpolation='nearest', aspect='auto')

    axs[0].set_title('Data 1')
    axs[1].set_title('Data 2')
    axs[2].set_title('Data 3')

    plt.show()"""

def calculate_error(original_matrix,input_matrix):
    #smallest one is better
    rmse = np.sqrt(np.mean((original_matrix - input_matrix) ** 2))

    return rmse

def output_function(input_matrix):
    file_names=os.listdir("data/result")

    if not file_names:
        print("---NO DATA WAS FOUND!---")
    else:
        matrices_dict={}
        len=len(file_names)
        for i in range(len):
            file_index=-180
            file_name = str(file_index)+".xlsx"
            wb = openpyxl.load_workbook('data/result/' + file_name)
            sheet = wb.active
            temp = []
            for row in sheet.iter_rows(min_row=1, max_row=4, min_col=1, max_col=141, values_only=True):
                temp.append(row)
            array = np.array(temp)
            error=calculate_error(array,input_matrix)
            key_name=str(file_index)

            matrices_dict[key_name]=(array,error)

            file_index=file_index+10

        sorted_error_values = sorted([(key, value[1]) for key, value in matrices_dict.items()], key=lambda x: x[1])
        min_key = int(sorted_error_values[0][0])


        print("Minimum error found in matrix: ",min_key)
        #error valueların grafini çıkarabilirsin
        #visulasie

        #hataların grafiğini görmek için aşağıdaki alanı yorum satırından çıkarın
        data = sorted_error_values
        x = [d[0] for d in data]
        y = [d[1] for d in data]

        fig, ax = plt.subplots(figsize=(15, 10))

        ax.plot(x, y, marker='o')
        ax.set_title('Line Graph')
        ax.set_xlabel('Matrix with Error')
        ax.set_ylabel('Value')
        ax.grid(True)

        plt.show()


        fig, axs = plt.subplots(nrows=1, ncols=2, figsize=(20, 5), sharex=True)

        detected_array=matrices_dict[str(min_key)][0]

        axs[0].imshow(detected_array, cmap='viridis', interpolation='nearest', aspect='auto')
        axs[1].imshow(input_matrix, cmap='viridis', interpolation='nearest', aspect='auto')

        axs[0].set_title('Detected Matrix')
        axs[1].set_title('Original Matrix')

        plt.show()
        return min_key


if __name__ == "__main__":
    DetectWhichPort() # USBports list is modified and sorted by sensor locations. (1, 2, 3, 4)
    p1 = Process(target=MyProc1, args=(USBports[0],))
    p2 = Process(target=MyProc2, args=(USBports[1],))
    p3 = Process(target=MyProc3, args=(USBports[2],))
    p4 = Process(target=MyProc4, args=(USBports[3],))

    p1.start()
    p2.start()
    p3.start()
    p4.start()

    arr1=[]
    arr2=[]
    arr3=[]
    arr4=[]



    x = input("harf gir > ")
    while x != "w":
        x = input("harf gir > ")
    
    for i in range(q1.qsize()):
        arr1.append(float(q1.get()))
        arr1_n=np.array(arr1)
    print("\n")
    for i in range(q2.qsize()):
        arr2.append(float(q2.get()))
        arr2_n=np.array(arr2)
    print("\n")
    for i in range(q3.qsize()):
        arr3.append(float(q3.get()))
        arr3_n=np.array(arr3)
    print("\n")
    for i in range(q4.qsize()):
        arr4.append(float(q4.get()))
        arr4_n=np.array(arr4)
    print("\n")

    input_matrix=np.concatenate(([arr1_n],[arr2_n],[arr3_n],[arr4_n]),axis=0)
    #control matrices sizes before concatenating. Zero matrix oluşturup ona atayabilrisin
    sel=int(input("input 1 for preparing data, press 2 for get a result"))
    if sel==1:
        data_preparetion(input_matrix)
        calculate_average()

    elif sel==2:
        output_function(input_matrix)
        #kod 4e 3 lük matrix için yazıldı onları düzenle. sonda transpose alındı daha iyi bir görüntü için ona da bir bak. ayrıca burası karşılaştırma  
        #için burayı da yazmayı tamamla. for döngüsü içindde results klasörü içindeki ortalamaları alıp karşılaştırma yaaabilirsin
        #genel bir kontrol yap ve kodu optimize et
        #ayrıca bu kısmı kaldırmak isteyebilirsin kod otonom bir araçta olacağı için kullanıcıdan input istemek mantıklı değil bunu iki ayrı kod halinde yapabilriz
        #üst tarafta sadece 3 tane matrix için yapıldı onu daha genel yap

        #sel==2 selden input almazsan direkt hesaba girecek
    else:
        print("invalid input")