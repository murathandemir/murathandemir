from multiprocessing import Process
import multiprocessing
import time
import serial
import RPi.GPIO as gpio

### ANGLE DETECT IMPORT BEGIN ###
import numpy as np
import matplotlib.pyplot as plt
import math as m
import pandas as pd
import os
import openpyxl
### ANGLE DETECT IMPORT END ###

DESCRIBE_YOURSELF_FLAG = "0".encode('utf-8')
MODE_SELECT = "2".encode('utf-8') # 1 for auto, 2 for manual.
IFFT_SELECT = "B".encode('utf-8') # A for enabled, B for disabled.
SEARCH_SEC = "8".encode('utf-8') # complement of 10.
SEARCH_SEC_INT = 2
BANDWIDTH_SELECT = "Z".encode('utf-8') # default is 3. for 5,7,9,11,13 > Y,X,W,V,U respectively.
DATAS_PER_BAND = 75
WAIT_FOR_POOL = 15
ANGLE_TRAIN_SELECTION = 1 # 1 for train mode, 2 for detection mode.
USEFUL_PRODUCT = DATAS_PER_BAND*SEARCH_SEC_INT
permissionFlag = [0,0,0,0]

# setting up the trigger.
EXTERNAL_TRIGGER_PIN = 20
gpio.setmode(gpio.BCM)
gpio.setup(EXTERNAL_TRIGGER_PIN, gpio.OUT)
gpio.setwarnings(False)

q1, q2, q3, q4 = multiprocessing.Queue(), multiprocessing.Queue(), multiprocessing.Queue(), multiprocessing.Queue() # Queues for processes.

USBports = [serial.Serial(), serial.Serial(), serial.Serial(), serial.Serial()]
UARTport = serial.Serial("/dev/ttyS0", baudrate=115200, timeout=0)
def DetectWhichPort():
    global USBports
    for i in range(0, 4):
        try:
            port = serial.Serial("/dev/ttyACM"+str(i), baudrate=115200, timeout=0)
            port.write(DESCRIBE_YOURSELF_FLAG)
            data = port.readline()
            while data == b'':
                data = port.readline()
            queue = int(data.decode('utf-8'))
            USBports[queue-1] = port
        except:
            None

def Config(port):
    port.write(MODE_SELECT)
    time.sleep(0.1)
    port.write(IFFT_SELECT)
    time.sleep(0.1)
    port.write(SEARCH_SEC)
    time.sleep(0.1)
    port.write(BANDWIDTH_SELECT)
    time.sleep(0.1)
    return 1

def MyProc1(port):
    Config(port)
    global q1, permissionFlag
    while 1:
        if permissionFlag[0]:
            try:
                data = port.readline()
                while data == b'':
                    data = port.readline()
                q1.put(data.decode('utf-8'))
                if q1.qsize() == (USEFUL_PRODUCT):
                    permissionFlag[0] = 0
            except:
                None

def MyProc2(port):
    Config(port)
    global q2, permissionFlag
    while 1:
        if permissionFlag[1]:
            try:
                data = port.readline()
                while data == b'':
                    data = port.readline()
                q2.put(data.decode('utf-8'))
                if q2.qsize() == (USEFUL_PRODUCT):
                    permissionFlag[1] = 0
            except:
                None

def MyProc3(port):
    Config(port)
    global q3, permissionFlag
    while 1:
        if permissionFlag[2]:
            try:
                data = port.readline()
                while data == b'':
                    data = port.readline()
                q3.put(data.decode('utf-8'))
                if q3.qsize() == (USEFUL_PRODUCT):
                    permissionFlag[2] = 0
            except:
                None    

def MyProc4(port):
    Config(port)
    global q4, permissionFlag
    while 1:
        if permissionFlag[3]:
            try:
                data = port.readline()
                while data == b'':
                    data = port.readline()
                q4.put(data.decode('utf-8'))
                if q4.qsize() == (USEFUL_PRODUCT):
                    permissionFlag[3] = 0
            except:
                None
        
def Trigger():
    global permissionFlag
    while 1:
        if permissionFlag[0] == 0 and permissionFlag[1] == 0 and permissionFlag[2] == 0 and permissionFlag[3] == 0:
            permissionFlag = [1,1,1,1]
            gpio.output(EXTERNAL_TRIGGER_PIN, gpio.HIGH)
            return 1
        else:
            continue

############### ANGLE DETECTION SIDE BEGIN ###############

def data_preparetion(input_matrix):
    print("Starting to data preparation!\n")

    while True:


#        print("2. On derece")
#        print("3. Eksi on derece")
#        print("0. Çıkış")


        sel = int(input("Please enter the degree: "))


        if sel == -180:
            max_number = 0
            if not os.path.exists("data/-180"):
                os.mkdir("data/-180")

            df = pd.DataFrame(input_matrix)
            file_names = os.listdir("data/-180")

            if not file_names:
                file_name='0.xlsx'

            else:
                for name in file_names:
                    if name.endswith(".xlsx"):
                        number = int(name[:-5])
                        if number > max_number:
                            max_number = number
                file_name = str(max_number+1) + ".xlsx"

            writer = pd.ExcelWriter('data/-180/'+file_name, engine='xlsxwriter')
            df.to_excel(writer, sheet_name='Sheet1', index=False, header=False)
            writer._save()

        elif sel == -170:
            max_number = 0
            if not os.path.exists("data/-170"):
                os.mkdir("data/-180")

            df = pd.DataFrame(input_matrix)
            file_names = os.listdir("data/-170")

            if not file_names:
                file_name = '0.xlsx'

            else:
                for name in file_names:
                    if name.endswith(".xlsx"):
                        number = int(name[:-5])
                        if number > max_number:
                            max_number = number
                file_name = str(max_number + 1) + ".xlsx"

            writer = pd.ExcelWriter('data/-180/' + file_name, engine='xlsxwriter')
            df.to_excel(writer, sheet_name='Sheet1', index=False, header=False)

            writer._save()

        elif sel == -160:
            max_number = 0
            if not os.path.exists("data/-160"):
                os.mkdir("data/-160")

            df = pd.DataFrame(input_matrix)
            file_names = os.listdir("data/-160")

            if not file_names:
                file_name='0.xlsx'

            else:
                for name in file_names:
                    if name.endswith(".xlsx"):
                        number = int(name[:-5])
                        if number > max_number:
                            max_number = number
                file_name = str(max_number+1) + ".xlsx"

            writer = pd.ExcelWriter('data/-160/'+file_name, engine='xlsxwriter')
            df.to_excel(writer, sheet_name='Sheet1', index=False, header=False)
            writer._save()

        elif sel == -150:
            max_number = 0
            if not os.path.exists("data/-150"):
                os.mkdir("data/-150")

            df = pd.DataFrame(input_matrix)
            file_names = os.listdir("data/-150")

            if not file_names:
                file_name = '0.xlsx'

            else:
                for name in file_names:
                    if name.endswith(".xlsx"):
                        number = int(name[:-5])
                        if number > max_number:
                            max_number = number
                file_name = str(max_number + 1) + ".xlsx"

            writer = pd.ExcelWriter('data/-150/' + file_name, engine='xlsxwriter')
            df.to_excel(writer, sheet_name='Sheet1', index=False, header=False)
            writer._save()

        elif sel == -140:
            max_number = 0
            if not os.path.exists("data/-140"):
                os.mkdir("data/-140")

            df = pd.DataFrame(input_matrix)
            file_names = os.listdir("data/-140")

            if not file_names:
                file_name = '0.xlsx'

            else:
                for name in file_names:
                    if name.endswith(".xlsx"):
                        number = int(name[:-5])
                        if number > max_number:
                            max_number = number
                file_name = str(max_number + 1) + ".xlsx"

            writer = pd.ExcelWriter('data/-140/' + file_name, engine='xlsxwriter')
            df.to_excel(writer, sheet_name='Sheet1', index=False, header=False)
            writer._save()

        elif sel == -130:
            max_number = 0
            if not os.path.exists("data/-130"):
                os.mkdir("data/-130")

            df = pd.DataFrame(input_matrix)
            file_names = os.listdir("data/-130")

            if not file_names:
                file_name = '0.xlsx'

            else:
                for name in file_names:
                    if name.endswith(".xlsx"):
                        number = int(name[:-5])
                        if number > max_number:
                            max_number = number
                file_name = str(max_number + 1) + ".xlsx"

            writer = pd.ExcelWriter('data/-130/' + file_name, engine='xlsxwriter')
            df.to_excel(writer, sheet_name='Sheet1', index=False, header=False)
            writer._save()

        elif sel == -120:
            max_number = 0
            if not os.path.exists("data/-120"):
                os.mkdir("data/-120")

            df = pd.DataFrame(input_matrix)
            file_names = os.listdir("data/-120")

            if not file_names:
                file_name = '0.xlsx'

            else:
                for name in file_names:
                    if name.endswith(".xlsx"):
                        number = int(name[:-5])
                        if number > max_number:
                            max_number = number
                file_name = str(max_number + 1) + ".xlsx"

            writer = pd.ExcelWriter('data/-120/' + file_name, engine='xlsxwriter')
            df.to_excel(writer, sheet_name='Sheet1', index=False, header=False)
            writer._save()

        elif sel == -110:
            max_number = 0
            if not os.path.exists("data/-110"):
                os.mkdir("data/-110")

            df = pd.DataFrame(input_matrix)
            file_names = os.listdir("data/-110")

            if not file_names:
                file_name = '0.xlsx'

            else:
                for name in file_names:
                    if name.endswith(".xlsx"):
                        number = int(name[:-5])
                        if number > max_number:
                            max_number = number
                file_name = str(max_number + 1) + ".xlsx"

            writer = pd.ExcelWriter('data/-110/' + file_name, engine='xlsxwriter')
            df.to_excel(writer, sheet_name='Sheet1', index=False, header=False)
            writer._save()

        elif sel == -100:
            max_number = 0
            if not os.path.exists("data/-100"):
                os.mkdir("data/-100")

            df = pd.DataFrame(input_matrix)
            file_names = os.listdir("data/-100")

            if not file_names:
                file_name = '0.xlsx'

            else:
                for name in file_names:
                    if name.endswith(".xlsx"):
                        number = int(name[:-5])
                        if number > max_number:
                            max_number = number
                file_name = str(max_number + 1) + ".xlsx"

            writer = pd.ExcelWriter('data/-100/' + file_name, engine='xlsxwriter')
            df.to_excel(writer, sheet_name='Sheet1', index=False, header=False)
            writer._save()

        elif sel == -90:
            max_number = 0
            if not os.path.exists("data/-90"):
                os.mkdir("data/-90")

            df = pd.DataFrame(input_matrix)
            file_names = os.listdir("data/-90")

            if not file_names:
                file_name = '0.xlsx'

            else:
                for name in file_names:
                    if name.endswith(".xlsx"):
                        number = int(name[:-5])
                        if number > max_number:
                            max_number = number
                file_name = str(max_number + 1) + ".xlsx"

            writer = pd.ExcelWriter('data/-90/' + file_name, engine='xlsxwriter')
            df.to_excel(writer, sheet_name='Sheet1', index=False, header=False)
            writer._save()

        elif sel == -80:
            max_number = 0
            if not os.path.exists("data/-80"):
                os.mkdir("data/-80")

            df = pd.DataFrame(input_matrix)
            file_names = os.listdir("data/-80")

            if not file_names:
                file_name = '0.xlsx'

            else:
                for name in file_names:
                    if name.endswith(".xlsx"):
                        number = int(name[:-5])
                        if number > max_number:
                            max_number = number
                file_name = str(max_number + 1) + ".xlsx"

            writer = pd.ExcelWriter('data/-80/' + file_name, engine='xlsxwriter')
            df.to_excel(writer, sheet_name='Sheet1', index=False, header=False)
            writer._save()

        elif sel == -70:
            max_number = 0
            if not os.path.exists("data/-70"):
                os.mkdir("data/-70")

            df = pd.DataFrame(input_matrix)
            file_names = os.listdir("data/-70")

            if not file_names:
                file_name = '0.xlsx'

            else:
                for name in file_names:
                    if name.endswith(".xlsx"):
                        number = int(name[:-5])
                        if number > max_number:
                            max_number = number
                file_name = str(max_number + 1) + ".xlsx"

            writer = pd.ExcelWriter('data/-70/' + file_name, engine='xlsxwriter')
            df.to_excel(writer, sheet_name='Sheet1', index=False, header=False)
            writer._save()

        elif sel == -60:
            max_number = 0
            if not os.path.exists("data/-60"):
                os.mkdir("data/-60")

            df = pd.DataFrame(input_matrix)
            file_names = os.listdir("data/-60")

            if not file_names:
                file_name = '0.xlsx'

            else:
                for name in file_names:
                    if name.endswith(".xlsx"):
                        number = int(name[:-5])
                        if number > max_number:
                            max_number = number
                file_name = str(max_number + 1) + ".xlsx"

            writer = pd.ExcelWriter('data/-60/' + file_name, engine='xlsxwriter')
            df.to_excel(writer, sheet_name='Sheet1', index=False, header=False)
            writer._save()

        elif sel == -50:
            max_number = 0
            if not os.path.exists("data/-50"):
                os.mkdir("data/-50")

            df = pd.DataFrame(input_matrix)
            file_names = os.listdir("data/-50")

            if not file_names:
                file_name = '0.xlsx'

            else:
                for name in file_names:
                    if name.endswith(".xlsx"):
                        number = int(name[:-5])
                        if number > max_number:
                            max_number = number
                file_name = str(max_number + 1) + ".xlsx"

            writer = pd.ExcelWriter('data/-60/' + file_name, engine='xlsxwriter')
            df.to_excel(writer, sheet_name='Sheet1', index=False, header=False)
            writer._save()

        elif sel == -40:
            max_number = 0
            if not os.path.exists("data/-40"):
                os.mkdir("data/-40")

            df = pd.DataFrame(input_matrix)
            file_names = os.listdir("data/-40")

            if not file_names:
                file_name = '0.xlsx'

            else:
                for name in file_names:
                    if name.endswith(".xlsx"):
                        number = int(name[:-5])
                        if number > max_number:
                            max_number = number
                file_name = str(max_number + 1) + ".xlsx"

            writer = pd.ExcelWriter('data/-40/' + file_name, engine='xlsxwriter')
            df.to_excel(writer, sheet_name='Sheet1', index=False, header=False)
            writer._save()

        elif sel == -30:
            max_number = 0
            if not os.path.exists("data/-30"):
                os.mkdir("data/-30")

            df = pd.DataFrame(input_matrix)
            file_names = os.listdir("data/-30")

            if not file_names:
                file_name = '0.xlsx'

            else:
                for name in file_names:
                    if name.endswith(".xlsx"):
                        number = int(name[:-5])
                        if number > max_number:
                            max_number = number
                file_name = str(max_number + 1) + ".xlsx"

            writer = pd.ExcelWriter('data/-30/' + file_name, engine='xlsxwriter')
            df.to_excel(writer, sheet_name='Sheet1', index=False, header=False)
            writer._save()

        elif sel == -20:
            max_number = 0
            if not os.path.exists("data/-20"):
                os.mkdir("data/-20")

            df = pd.DataFrame(input_matrix)
            file_names = os.listdir("data/-20")

            if not file_names:
                file_name = '0.xlsx'

            else:
                for name in file_names:
                    if name.endswith(".xlsx"):
                        number = int(name[:-5])
                        if number > max_number:
                            max_number = number
                file_name = str(max_number + 1) + ".xlsx"

            writer = pd.ExcelWriter('data/-20/' + file_name, engine='xlsxwriter')
            df.to_excel(writer, sheet_name='Sheet1', index=False, header=False)
            writer._save()

        elif sel == -10:
            max_number = 0
            if not os.path.exists("data/-10"):
                os.mkdir("data/-10")

            df = pd.DataFrame(input_matrix)
            file_names = os.listdir("data/-10")

            if not file_names:
                file_name = '0.xlsx'

            else:
                for name in file_names:
                    if name.endswith(".xlsx"):
                        number = int(name[:-5])
                        if number > max_number:
                            max_number = number
                file_name = str(max_number + 1) + ".xlsx"

            writer = pd.ExcelWriter('data/-10/' + file_name, engine='xlsxwriter')
            df.to_excel(writer, sheet_name='Sheet1', index=False, header=False)
            writer._save()

        elif sel == 0:
            max_number = 0
            if not os.path.exists("data/0"):
                os.mkdir("data/0")

            df = pd.DataFrame(input_matrix)
            file_names = os.listdir("data/0")

            if not file_names:
                file_name = '0.xlsx'

            else:
                for name in file_names:
                    if name.endswith(".xlsx"):
                        number = int(name[:-5])
                        if number > max_number:
                            max_number = number
                file_name = str(max_number + 1) + ".xlsx"

            writer = pd.ExcelWriter('data/0/' + file_name, engine='xlsxwriter')
            df.to_excel(writer, sheet_name='Sheet1', index=False, header=False)
            writer._save()

        elif sel == 180:
            max_number = 0
            if not os.path.exists("data/180"):
                os.mkdir("data/180")

            df = pd.DataFrame(input_matrix)
            file_names = os.listdir("data/180")

            if not file_names:
                file_name = '0.xlsx'

            else:
                for name in file_names:
                    if name.endswith(".xlsx"):
                        number = int(name[:-5])
                        if number > max_number:
                            max_number = number
                file_name = str(max_number + 1) + ".xlsx"

            writer = pd.ExcelWriter('data/180/' + file_name, engine='xlsxwriter')
            df.to_excel(writer, sheet_name='Sheet1', index=False, header=False)
            writer._save()

        elif sel == 170:
            max_number = 0
            if not os.path.exists("data/170"):
                os.mkdir("data/170")

            df = pd.DataFrame(input_matrix)
            file_names = os.listdir("data/170")

            if not file_names:
                file_name = '0.xlsx'

            else:
                for name in file_names:
                    if name.endswith(".xlsx"):
                        number = int(name[:-5])
                        if number > max_number:
                            max_number = number
                file_name = str(max_number + 1) + ".xlsx"

            writer = pd.ExcelWriter('data/170/' + file_name, engine='xlsxwriter')
            df.to_excel(writer, sheet_name='Sheet1', index=False, header=False)

            writer._save()

        elif sel == 160:
            max_number = 0
            if not os.path.exists("data/160"):
                os.mkdir("data/160")

            df = pd.DataFrame(input_matrix)
            file_names = os.listdir("data/160")

            if not file_names:
                file_name = '0.xlsx'

            else:
                for name in file_names:
                    if name.endswith(".xlsx"):
                        number = int(name[:-5])
                        if number > max_number:
                            max_number = number
                file_name = str(max_number + 1) + ".xlsx"

            writer = pd.ExcelWriter('data/160/' + file_name, engine='xlsxwriter')
            df.to_excel(writer, sheet_name='Sheet1', index=False, header=False)
            writer._save()

        elif sel == 150:
            max_number = 0
            if not os.path.exists("data/150"):
                os.mkdir("data/150")

            df = pd.DataFrame(input_matrix)
            file_names = os.listdir("data/150")

            if not file_names:
                file_name = '0.xlsx'

            else:
                for name in file_names:
                    if name.endswith(".xlsx"):
                        number = int(name[:-5])
                        if number > max_number:
                            max_number = number
                file_name = str(max_number + 1) + ".xlsx"

            writer = pd.ExcelWriter('data/150/' + file_name, engine='xlsxwriter')
            df.to_excel(writer, sheet_name='Sheet1', index=False, header=False)
            writer._save()

        elif sel == 140:
            max_number = 0
            if not os.path.exists("data/140"):
                os.mkdir("data/140")

            df = pd.DataFrame(input_matrix)
            file_names = os.listdir("data/140")

            if not file_names:
                file_name = '0.xlsx'

            else:
                for name in file_names:
                    if name.endswith(".xlsx"):
                        number = int(name[:-5])
                        if number > max_number:
                            max_number = number
                file_name = str(max_number + 1) + ".xlsx"

            writer = pd.ExcelWriter('data/140/' + file_name, engine='xlsxwriter')
            df.to_excel(writer, sheet_name='Sheet1', index=False, header=False)
            writer._save()

        elif sel == 130:
            max_number = 0
            if not os.path.exists("data/130"):
                os.mkdir("data/130")

            df = pd.DataFrame(input_matrix)
            file_names = os.listdir("data/130")

            if not file_names:
                file_name = '0.xlsx'

            else:
                for name in file_names:
                    if name.endswith(".xlsx"):
                        number = int(name[:-5])
                        if number > max_number:
                            max_number = number
                file_name = str(max_number + 1) + ".xlsx"

            writer = pd.ExcelWriter('data/130/' + file_name, engine='xlsxwriter')
            df.to_excel(writer, sheet_name='Sheet1', index=False, header=False)
            writer._save()

        elif sel == 120:
            max_number = 0
            if not os.path.exists("data/120"):
                os.mkdir("data/120")

            df = pd.DataFrame(input_matrix)
            file_names = os.listdir("data/120")

            if not file_names:
                file_name = '0.xlsx'

            else:
                for name in file_names:
                    if name.endswith(".xlsx"):
                        number = int(name[:-5])
                        if number > max_number:
                            max_number = number
                file_name = str(max_number + 1) + ".xlsx"

            writer = pd.ExcelWriter('data/120/' + file_name, engine='xlsxwriter')
            df.to_excel(writer, sheet_name='Sheet1', index=False, header=False)
            writer._save()

        elif sel == 110:
            max_number = 0
            if not os.path.exists("data/110"):
                os.mkdir("data/110")

            df = pd.DataFrame(input_matrix)
            file_names = os.listdir("data/110")

            if not file_names:
                file_name = '0.xlsx'

            else:
                for name in file_names:
                    if name.endswith(".xlsx"):
                        number = int(name[:-5])
                        if number > max_number:
                            max_number = number
                file_name = str(max_number + 1) + ".xlsx"

            writer = pd.ExcelWriter('data/110/' + file_name, engine='xlsxwriter')
            df.to_excel(writer, sheet_name='Sheet1', index=False, header=False)
            writer._save()

        elif sel == 100:
            max_number = 0
            if not os.path.exists("data/100"):
                os.mkdir("data/100")

            df = pd.DataFrame(input_matrix)
            file_names = os.listdir("data/100")

            if not file_names:
                file_name = '0.xlsx'

            else:
                for name in file_names:
                    if name.endswith(".xlsx"):
                        number = int(name[:-5])
                        if number > max_number:
                            max_number = number
                file_name = str(max_number + 1) + ".xlsx"

            writer = pd.ExcelWriter('data/100/' + file_name, engine='xlsxwriter')
            df.to_excel(writer, sheet_name='Sheet1', index=False, header=False)
            writer._save()

        elif sel == 90:
            max_number = 0
            if not os.path.exists("data/90"):
                os.mkdir("data/90")

            df = pd.DataFrame(input_matrix)
            file_names = os.listdir("data/90")

            if not file_names:
                file_name = '0.xlsx'

            else:
                for name in file_names:
                    if name.endswith(".xlsx"):
                        number = int(name[:-5])
                        if number > max_number:
                            max_number = number
                file_name = str(max_number + 1) + ".xlsx"

            writer = pd.ExcelWriter('data/90/' + file_name, engine='xlsxwriter')
            df.to_excel(writer, sheet_name='Sheet1', index=False, header=False)
            writer._save()

        elif sel == 80:
            max_number = 0
            if not os.path.exists("data/80"):
                os.mkdir("data/80")

            df = pd.DataFrame(input_matrix)
            file_names = os.listdir("data/80")

            if not file_names:
                file_name = '0.xlsx'

            else:
                for name in file_names:
                    if name.endswith(".xlsx"):
                        number = int(name[:-5])
                        if number > max_number:
                            max_number = number
                file_name = str(max_number + 1) + ".xlsx"

            writer = pd.ExcelWriter('data/80/' + file_name, engine='xlsxwriter')
            df.to_excel(writer, sheet_name='Sheet1', index=False, header=False)
            writer._save()

        elif sel == 70:
            max_number = 0
            if not os.path.exists("data/70"):
                os.mkdir("data/70")

            df = pd.DataFrame(input_matrix)
            file_names = os.listdir("data/70")

            if not file_names:
                file_name = '0.xlsx'

            else:
                for name in file_names:
                    if name.endswith(".xlsx"):
                        number = int(name[:-5])
                        if number > max_number:
                            max_number = number
                file_name = str(max_number + 1) + ".xlsx"

            writer = pd.ExcelWriter('data/70/' + file_name, engine='xlsxwriter')
            df.to_excel(writer, sheet_name='Sheet1', index=False, header=False)
            writer._save()

        elif sel == 60:
            max_number = 0
            if not os.path.exists("data/60"):
                os.mkdir("data/60")

            df = pd.DataFrame(input_matrix)
            file_names = os.listdir("data/60")

            if not file_names:
                file_name = '0.xlsx'

            else:
                for name in file_names:
                    if name.endswith(".xlsx"):
                        number = int(name[:-5])
                        if number > max_number:
                            max_number = number
                file_name = str(max_number + 1) + ".xlsx"

            writer = pd.ExcelWriter('data/60/' + file_name, engine='xlsxwriter')
            df.to_excel(writer, sheet_name='Sheet1', index=False, header=False)
            writer._save()

        elif sel == 50:
            max_number = 0
            if not os.path.exists("data/50"):
                os.mkdir("data/50")

            df = pd.DataFrame(input_matrix)
            file_names = os.listdir("data/50")

            if not file_names:
                file_name = '0.xlsx'

            else:
                for name in file_names:
                    if name.endswith(".xlsx"):
                        number = int(name[:-5])
                        if number > max_number:
                            max_number = number
                file_name = str(max_number + 1) + ".xlsx"

            writer = pd.ExcelWriter('data/50/' + file_name, engine='xlsxwriter')
            df.to_excel(writer, sheet_name='Sheet1', index=False, header=False)
            writer._save()

        elif sel == 40:
            max_number = 0
            if not os.path.exists("data/40"):
                os.mkdir("data/40")

            df = pd.DataFrame(input_matrix)
            file_names = os.listdir("data/40")

            if not file_names:
                file_name = '0.xlsx'

            else:
                for name in file_names:
                    if name.endswith(".xlsx"):
                        number = int(name[:-5])
                        if number > max_number:
                            max_number = number
                file_name = str(max_number + 1) + ".xlsx"

            writer = pd.ExcelWriter('data/40/' + file_name, engine='xlsxwriter')
            df.to_excel(writer, sheet_name='Sheet1', index=False, header=False)
            writer._save()

        elif sel == 30:
            max_number = 0
            if not os.path.exists("data/30"):
                os.mkdir("data/30")

            df = pd.DataFrame(input_matrix)
            file_names = os.listdir("data/30")

            if not file_names:
                file_name = '0.xlsx'

            else:
                for name in file_names:
                    if name.endswith(".xlsx"):
                        number = int(name[:-5])
                        if number > max_number:
                            max_number = number
                file_name = str(max_number + 1) + ".xlsx"

            writer = pd.ExcelWriter('data/30/' + file_name, engine='xlsxwriter')
            df.to_excel(writer, sheet_name='Sheet1', index=False, header=False)
            writer._save()

        elif sel == 20:
            max_number = 0
            if not os.path.exists("data/20"):
                os.mkdir("data/20")

            df = pd.DataFrame(input_matrix)
            file_names = os.listdir("data/20")

            if not file_names:
                file_name = '0.xlsx'

            else:
                for name in file_names:
                    if name.endswith(".xlsx"):
                        number = int(name[:-5])
                        if number > max_number:
                            max_number = number
                file_name = str(max_number + 1) + ".xlsx"

            writer = pd.ExcelWriter('data/20/' + file_name, engine='xlsxwriter')
            df.to_excel(writer, sheet_name='Sheet1', index=False, header=False)
            writer._save()

        elif sel == 10:
            max_number = 0
            if not os.path.exists("data/10"):
                os.mkdir("data/10")

            df = pd.DataFrame(input_matrix)
            file_names = os.listdir("data/10")

            if not file_names:
                file_name = '0.xlsx'

            else:
                for name in file_names:
                    if name.endswith(".xlsx"):
                        number = int(name[:-5])
                        if number > max_number:
                            max_number = number
                file_name = str(max_number + 1) + ".xlsx"

            writer = pd.ExcelWriter('data/10/' + file_name, engine='xlsxwriter')
            df.to_excel(writer, sheet_name='Sheet1', index=False, header=False)
            writer._save()

        elif sel == 52:
            print("Programdan çıkılıyor...")
            break
        else:
            print("Geçersiz seçim. Lütfen tekrar deneyin.")

def calculate_average():
    if not os.path.exists("data/results"):
        os.mkdir("data/results")

    matrix_size=(4,int(USEFUL_PRODUCT))
    for i in range(len(os.listdir('data'))):
        dir_index=-180
        file_index=0
        sum=np.zeros(matrix_size)
        dir="data/"+str(dir_index)
        file_names=os.listdir(dir)
        len=len(file_names)

        for j in range(len):
            file_name=str(file_index)+".xlsx"
            wb=openpyxl.load_workbook("data/"+str(dir_index)+"/"+file_name)
            sheet=wb.active
            temp=[]
            for row in sheet.iter_rows(min_row=1,max_row=4,min_col=1,max_col=int(USEFUL_PRODUCT),values_only=True):
                temp.append(row)
            array=np.array(temp)
            sum=array+sum

        average=sum/len


        df = pd.DataFrame(average)
        writer = pd.ExcelWriter("data/results/"+str(dir_index)+".xlsx", engine='xlsxwriter')
        df.to_excel(writer, sheet_name='Sheet1', index=False, header=False)
        writer._save()

        dir_index=dir_index+10

    print("---Average matrices calculates and saved into 'data/results'---")



    """
    fig, axs = plt.subplots(1, 3, figsize=(10, 4))

    axs[0].imshow(average_1,cmap='viridis',interpolation='nearest', aspect='auto')
    axs[1].imshow(average_2,cmap='viridis',interpolation='nearest', aspect='auto')
    axs[2].imshow(average_3,cmap='viridis',interpolation='nearest', aspect='auto')

    axs[0].set_title('Data 1')
    axs[1].set_title('Data 2')
    axs[2].set_title('Data 3')

    plt.show()"""

def calculate_error(original_matrix,input_matrix):
    #smallest one is better
    rmse = np.sqrt(np.mean((original_matrix - input_matrix) ** 2))

    return rmse

def output_function(input_matrix):
    file_names=os.listdir("data/result")

    if not file_names:
        print("---NO DATA WAS FOUND!---")
    else:
        matrices_dict={}
        len=len(file_names)
        for i in range(len):
            file_index=-180
            file_name = str(file_index)+".xlsx"
            wb = openpyxl.load_workbook('data/result/' + file_name)
            sheet = wb.active
            temp = []
            for row in sheet.iter_rows(min_row=1, max_row=4, min_col=1, max_col=int(USEFUL_PRODUCT), values_only=True):
                temp.append(row)
            array = np.array(temp)
            error=calculate_error(array,input_matrix)
            key_name=str(file_index)

            matrices_dict[key_name]=(array,error)

            file_index=file_index+10

        sorted_error_values = sorted([(key, value[1]) for key, value in matrices_dict.items()], key=lambda x: x[1])
        min_key = int(sorted_error_values[0][0])


        print("Minimum error found in matrix: ",min_key)
        #error valueların grafini çıkarabilirsin
        #visulasie

        #hataların grafiğini görmek için aşağıdaki alanı yorum satırından çıkarın
        data = sorted_error_values
        x = [d[0] for d in data]
        y = [d[1] for d in data]

        fig, ax = plt.subplots(figsize=(15, 10))

        ax.plot(x, y, marker='o')
        ax.set_title('Line Graph')
        ax.set_xlabel('Matrix with Error')
        ax.set_ylabel('Value')
        ax.grid(True)

        plt.show()


        fig, axs = plt.subplots(nrows=1, ncols=2, figsize=(20, 5), sharex=True)

        detected_array=matrices_dict[str(min_key)][0]

        axs[0].imshow(detected_array, cmap='viridis', interpolation='nearest', aspect='auto')
        axs[1].imshow(input_matrix, cmap='viridis', interpolation='nearest', aspect='auto')

        axs[0].set_title('Detected Matrix')
        axs[1].set_title('Original Matrix')

        plt.show()
        return min_key

############### ANGLE DETECTION SIDE END ###############

if __name__ == "__main__":

    while 1:
        UARTport.write("h".encode('utf-8'))
        isOpen = UARTport.read()
        if isOpen.decode('utf-8') == "f":
            time.sleep(WAIT_FOR_POOL)
            break

    try:
        DetectWhichPort() # USBports list is modified and sorted by sensor locations. (1, 2, 3, 4)
    except:
        print("SENSORS COULD NOT BE IDENTIFIED. SCRIPT ABORTED.")
        exit()

    try:
        f1 = open("data1.txt", "a")
        f2 = open("data2.txt", "a")
        f3 = open("data3.txt", "a")
        f4 = open("data4.txt", "a")
    except:
        None
    
    p1 = Process(target=MyProc1, args=(USBports[0],))
    p2 = Process(target=MyProc2, args=(USBports[1],))
    p3 = Process(target=MyProc3, args=(USBports[2],))
    p4 = Process(target=MyProc4, args=(USBports[3],))

    p1.start()
    p2.start()
    p3.start()
    p4.start()

    firstSonarList = []
    secondSonarList = []
    thirdSonarList = []
    fourthSonarList = []

    Trigger() # Hidrofondan veriyi okumaya başla.

    while 1:
        waitSeconds = int(SEARCH_SEC_INT * 1.6)
        time.sleep(waitSeconds) # wait till sensors sense the environment.
        permissionFlag = [0,0,0,0] # dont listen just a bit
        for i in range(q1.qsize()):
            firstSonarList.append(float(q1.get()))
            arr1_n = np.array(firstSonarList)
        for i in range(q2.qsize()):
            secondSonarList.append(float(q2.get()))
            arr2_n = np.array(secondSonarList)
        for i in range(q3.qsize()):
            thirdSonarList.append(float(q3.get()))
            arr3_n = np.array(thirdSonarList)
        for i in range(q4.qsize()):
            fourthSonarList.append(float(q4.get()))
            arr4_n = np.array(fourthSonarList)

        while (len(arr1_n != int(USEFUL_PRODUCT))):
            if(len(arr1_n) > int(USEFUL_PRODUCT)):
                arr1_n = np.delete(arr1_n, -1)
            else:
                arr1_n = np.append(arr1_n, 0)

        while (len(arr2_n != int(USEFUL_PRODUCT))):
            if(len(arr2_n) > int(USEFUL_PRODUCT)):
                arr2_n = np.delete(arr2_n, -1)
            else:
                arr2_n = np.append(arr2_n, 0)

        while (len(arr3_n != int(USEFUL_PRODUCT))):
            if(len(arr3_n) > int(USEFUL_PRODUCT)):
                arr3_n = np.delete(arr3_n, -1)
            else:
                arr3_n = np.append(arr3_n, 0)

        while (len(arr4_n != int(USEFUL_PRODUCT))):
            if(len(arr4_n) > int(USEFUL_PRODUCT)):
                arr4_n = np.delete(arr4_n, -1)
            else:
                arr4_n = np.append(arr4_n, 0)
        

        for i in range(len(firstSonarList)):
            f1.write(str((i+1)) + str(firstSonarList[i]) + "\n")

        for i in range(len(secondSonarList)):
            f2.write(str((i+1)) + str(secondSonarList[i]) + "\n")

        for i in range(len(thirdSonarList)):
            f3.write(str((i+1)) + str(thirdSonarList[i]) + "\n")

        for i in range(len(fourthSonarList)):
            f4.write(str((i+1)) + str(fourthSonarList[i]) + "\n")

        USBports[0].reset_input_buffer()
        USBports[1].reset_input_buffer()
        USBports[2].reset_input_buffer()
        USBports[3].reset_input_buffer()

        ############# AÇIYI HESAPLA BAŞLANGIÇ #############

        input_matrix = np.concatenate(([arr1_n],[arr2_n],[arr3_n],[arr4_n]),axis=0)
        sel = ANGLE_TRAIN_SELECTION
        tempAngle = 0
        if sel==1:
            data_preparetion(input_matrix)
            calculate_average()

        elif sel==2:
            tempAngle = output_function(input_matrix)

        else:
            print("invalid input")

        ############# AÇIYI HESAPLA BİTİŞ #############

        # AÇI GÖNDER

        """
        choose = "-"
        while 1:
            is_g = UARTport.readline()
            if is_g.decode('utf-8') == "g":
                choose = "g"
                break
        """

        choose = input("tamam ise 0, devam ise g > ") # it is just for debug. default is 1.
        if choose == "0":
            p1.terminate() # kill all processes
            p2.terminate()
            p3.terminate()
            p4.terminate()
            USBports[0].close() # close all ports
            USBports[1].close()
            USBports[2].close()
            USBports[3].close()
            UARTport.close()
            f1.close() # close all files
            f2.close()
            f3.close()
            f4.close()
            exit()
        elif choose == "g":
            angle = tempAngle
            UARTport.write(angle.encode('utf-8')) 
            Trigger()
            time.sleep(0.1)
            gpio.output(EXTERNAL_TRIGGER_PIN, gpio.LOW)
            choose = "-"
            continue
        else:
            None